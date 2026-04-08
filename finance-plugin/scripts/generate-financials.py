#!/usr/bin/env python3
# /// script
# requires-python = ">=3.9"
# dependencies = ["openpyxl"]
# ///
"""
generate-financials.py — Generate Excel workbook and print-ready HTML from finance-data.json.

Produces:
  - {period}-financial-pack.xlsx    — Multi-sheet workbook (ledger + statements + summaries)
  - {period}-financial-statements.html — Print-optimised HTML (Ctrl+P → PDF)

Usage:
  python3 finance-plugin/scripts/generate-financials.py --period FY2025 --output all
  python3 finance-plugin/scripts/generate-financials.py --period FY2025 --output excel
  python3 finance-plugin/scripts/generate-financials.py --period FY2025 --output html

Output: finance/tax/{period}/
"""

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime
from html import escape
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: 'openpyxl' not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ─── Configuration ────────────────────────────────────────────────────────────

FINANCE_DATA_PATH = Path("finance/finance-data.json")
BALANCES_PATH = Path("finance/balances.json")
MANUAL_DEDUCTIONS_FILENAME = "manual-deductions.json"

# FY period parsing: FY2025 = 1 Jul 2024 – 30 Jun 2025
def parse_period(period: str):
    if period.startswith("FY"):
        year = int(period[2:])
        return f"{year - 1}-07-01", f"{year}-06-30", period
    else:
        raise ValueError(f"Unsupported period format: {period}. Use FY2025.")


# ─── Design tokens ────────────────────────────────────────────────────────────

BRAND_PRIMARY = "#7DFF00"
BRAND_DARK = "#0f1825"
BRAND_GREY = "#64748b"
BRAND_LIGHT = "#f7f9fc"
COMPANY_LOGO_URL = "{YOUR_LOGO_URL}"

# Excel colors (no # prefix for openpyxl)
XL_DARK = "0f1825"
XL_LIME = "7DFF00"
XL_LIGHT = "f7f9fc"
XL_WHITE = "FFFFFF"
XL_BORDER_COLOR = "e2e8f0"


def fmt_aud(amount) -> str:
    if amount is None:
        return "N/A"
    return f"${amount:,.2f}"


def fmt_aud_int(amount) -> str:
    if amount is None:
        return "N/A"
    return f"${amount:,.0f}"


# ─── Data Loading ─────────────────────────────────────────────────────────────

def load_data():
    with open(FINANCE_DATA_PATH) as f:
        return json.load(f)


def load_balances():
    if BALANCES_PATH.exists():
        with open(BALANCES_PATH) as f:
            return json.load(f)
    return {"accounts": []}


def aud(t):
    """Get AUD amount for a transaction."""
    return t.get("amount_aud") or t.get("amount", 0)


def abs_aud(t):
    return abs(aud(t))


def filter_period(transactions, from_date, to_date):
    return [t for t in transactions if from_date <= (t.get("date") or "") <= to_date]


# ─── Calculations ─────────────────────────────────────────────────────────────

def load_manual_deductions(period):
    """Load manual deductions file if it exists for the given period."""
    path = Path(f"finance/tax/{period}/{MANUAL_DEDUCTIONS_FILENAME}")
    if path.exists():
        with open(path) as f:
            return json.load(f)
    return None


def calculate_statements(txns, balances, manual_deductions=None):
    """Calculate all financial statement data from transactions.

    Uses transaction_class to determine what counts as revenue vs expense:
      - "income"        → genuine external revenue (client deposits)
      - "expense"       → genuine business expenses (spend, fees, contractor payments)
      - "internal"      → FX conversions, wallet transfers, payment attempts (EXCLUDED from P&L)
      - "refund"        → failed payment refunds, card refunds (EXCLUDED from revenue, offset against expenses)
      - "amex_payment"  → paying your own Amex bill (EXCLUDED — internal movement)
    """
    # Revenue = only transactions classified as "income"
    revenue_txns = [t for t in txns if t.get("transaction_class") == "income"]
    # Expenses = only transactions classified as "expense"
    expense_txns = [t for t in txns if t.get("transaction_class") == "expense"]
    # Refunds offset against expenses (reduce total spend)
    refund_txns = [t for t in txns if t.get("transaction_class") == "refund"]
    # Internal and amex_payment are excluded from P&L entirely
    internal_txns = [t for t in txns if t.get("transaction_class") in ("internal", "amex_payment")]

    # Revenue
    revenue_total = sum(abs_aud(t) for t in revenue_txns)
    revenue_by_merchant = defaultdict(float)
    for t in revenue_txns:
        name = t.get("merchant_name_normalized") or t.get("description") or "Unknown"
        revenue_by_merchant[name] += abs_aud(t)

    # Revenue GST split
    revenue_gst_free = sum(abs_aud(t) for t in revenue_txns if t.get("gst_status") == "exempt")
    revenue_gst_incl = sum(abs_aud(t) for t in revenue_txns if t.get("gst_status") == "included")
    revenue_gst_collected = sum(t.get("gst_amount", 0) or 0 for t in revenue_txns if t.get("gst_status") == "included")

    # Expenses by category
    expense_by_cat = defaultdict(lambda: {"total": 0.0, "count": 0})
    for t in expense_txns:
        cat = t.get("category_id") or "uncategorized"
        expense_by_cat[cat]["total"] += abs_aud(t)
        expense_by_cat[cat]["count"] += 1

    # Manual deductions (non-bank expenses)
    manual_total = 0.0
    manual_by_cat = {}
    if manual_deductions:
        for d in manual_deductions.get("deductions", []):
            cat = d["category"]
            amt = d["amount"]
            manual_by_cat[cat] = amt
            manual_total += amt
            # Add to expense_by_cat for unified reporting
            expense_by_cat[cat]["total"] += amt
            expense_by_cat[cat]["count"] += 1

    # Refunds reduce expenses
    refund_total = sum(abs_aud(t) for t in refund_txns)

    cogs = expense_by_cat.pop("contractors", {"total": 0, "count": 0})["total"]
    uncat = expense_by_cat.pop("uncategorized", {"total": 0, "count": 0})["total"]
    opex = {k: v["total"] for k, v in expense_by_cat.items()}
    total_opex = sum(opex.values())
    gross_profit = revenue_total - cogs
    net_profit = gross_profit - total_opex - uncat + refund_total

    # Cash flow (uses same classification — manual deductions are non-cash)
    cash_received = revenue_total
    all_expenses = sum(abs_aud(t) for t in expense_txns)
    investing = sum(abs_aud(t) for t in expense_txns
                    if (t.get("category_id") or "") in ("office", "equipment", "computer") and abs_aud(t) > 300)
    operating_paid = all_expenses - investing

    # Balance
    aud_balance = 0
    for acc in balances.get("accounts", []):
        if acc.get("currency") == "AUD":
            aud_balance = acc.get("available_amount", 0)

    # Category summary
    cat_summary = defaultdict(lambda: {"count": 0, "total": 0.0})
    for t in txns:
        cat = t.get("category_id") or "uncategorized"
        cat_summary[cat]["count"] += 1
        cat_summary[cat]["total"] += abs_aud(t)

    # FX summary
    fx_summary = defaultdict(lambda: {"count": 0, "original_total": 0.0, "aud_total": 0.0, "rates": [], "sources": set()})
    for t in txns:
        cur = t.get("currency") or "AUD"
        fx_summary[cur]["count"] += 1
        fx_summary[cur]["original_total"] += abs(t.get("amount", 0))
        fx_summary[cur]["aud_total"] += abs_aud(t)
        fx = t.get("fx") or {}
        if fx.get("fx_rate"):
            fx_summary[cur]["rates"].append(fx["fx_rate"])
        if fx.get("fx_source"):
            fx_summary[cur]["sources"].add(fx["fx_source"])

    # Completeness
    categorized = sum(1 for t in txns if t.get("category_confidence") != "UNSET")
    has_aud_amt = sum(1 for t in txns if t.get("amount_aud") is not None)
    gst_classified = sum(1 for t in txns if t.get("gst_status") not in ("unknown", None))

    return {
        "revenue_total": revenue_total,
        "revenue_by_merchant": dict(revenue_by_merchant),
        "revenue_gst_free": revenue_gst_free,
        "revenue_gst_incl": revenue_gst_incl,
        "revenue_gst_collected": revenue_gst_collected,
        "cogs": cogs,
        "gross_profit": gross_profit,
        "opex": opex,
        "total_opex": total_opex,
        "uncategorized": uncat,
        "refund_total": refund_total,
        "net_profit": net_profit,
        "manual_deductions": manual_by_cat,
        "manual_total": manual_total,
        "cash_received": cash_received,
        "operating_paid": operating_paid,
        "investing": investing,
        "operating_net": cash_received - operating_paid + refund_total,
        "financing_net": 0,
        "net_change": cash_received - all_expenses + refund_total,
        "aud_balance": aud_balance,
        "cat_summary": dict(cat_summary),
        "fx_summary": dict(fx_summary),
        "total_txns": len(txns),
        "revenue_txn_count": len(revenue_txns),
        "expense_txn_count": len(expense_txns),
        "internal_txn_count": len(internal_txns),
        "refund_txn_count": len(refund_txns),
        "internal_total": sum(abs_aud(t) for t in internal_txns),
        "categorized": categorized,
        "has_aud": has_aud_amt,
        "gst_classified": gst_classified,
    }


# ─── Excel Generation ─────────────────────────────────────────────────────────

def xl_styles():
    """Return reusable Excel styles."""
    header_font = Font(name="Calibri", bold=True, color=XL_WHITE, size=11)
    header_fill = PatternFill(start_color=XL_DARK, end_color=XL_DARK, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", bold=True, size=10)
    title_font = Font(name="Calibri", bold=True, size=14, color=XL_DARK)
    subtitle_font = Font(name="Calibri", bold=True, size=11, color=XL_DARK)

    accent_fill = PatternFill(start_color=XL_LIGHT, end_color=XL_LIGHT, fill_type="solid")
    lime_fill = PatternFill(start_color=XL_LIME, end_color=XL_LIME, fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color=XL_BORDER_COLOR),
        right=Side(style="thin", color=XL_BORDER_COLOR),
        top=Side(style="thin", color=XL_BORDER_COLOR),
        bottom=Side(style="thin", color=XL_BORDER_COLOR),
    )

    return {
        "header_font": header_font, "header_fill": header_fill, "header_align": header_align,
        "data_font": data_font, "bold_font": bold_font, "title_font": title_font,
        "subtitle_font": subtitle_font, "accent_fill": accent_fill, "lime_fill": lime_fill,
        "thin_border": thin_border,
    }


def write_ledger_sheet(wb, txns, styles):
    """Sheet 1: Transaction Ledger — every transaction, one per row."""
    ws = wb.create_sheet("Transaction Ledger")

    headers = [
        "Date", "Posted Date", "Description", "Category", "Class",
        "Original Amount", "Original Currency", "AUD Amount", "FX Rate",
        "FX Source", "Direction", "Source", "GST Status", "GST Amount",
        "Transaction ID", "Airwallex ID", "Receipt", "Tags",
    ]

    # Write header
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["header_align"]
        cell.border = styles["thin_border"]

    # Write data (sorted by date ascending)
    sorted_txns = sorted(txns, key=lambda t: t.get("date", ""))
    for row_idx, t in enumerate(sorted_txns, 2):
        fx = t.get("fx") or {}
        receipt = t.get("receipt") or {}
        tags = t.get("tags") or []

        values = [
            t.get("date", ""),
            t.get("posted_date", ""),
            t.get("merchant_name") or t.get("description", ""),
            t.get("category_id", ""),
            t.get("transaction_class", ""),
            t.get("amount", 0),
            t.get("currency", "AUD"),
            t.get("amount_aud"),
            fx.get("fx_rate"),
            fx.get("fx_source", ""),
            t.get("direction", ""),
            t.get("source", ""),
            t.get("gst_status", ""),
            t.get("gst_amount"),
            t.get("transaction_id", ""),
            t.get("airwallex_id", ""),
            "Yes" if receipt.get("attached") else "No",
            ", ".join(tags) if tags else "",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = styles["data_font"]
            cell.border = styles["thin_border"]

    # Format currency columns
    for row in range(2, len(sorted_txns) + 2):
        for col in [6, 8, 14]:  # Original Amount, AUD Amount, GST Amount
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    # Auto-filter and freeze
    ws.auto_filter.ref = f"A1:R{len(sorted_txns) + 1}"
    ws.freeze_panes = "A2"

    # Column widths
    widths = [11, 11, 40, 15, 12, 14, 10, 14, 10, 10, 8, 20, 10, 12, 22, 22, 8, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_income_statement_sheet(wb, calc, period, styles):
    """Sheet 2: Income Statement (P&L)."""
    ws = wb.create_sheet("Income Statement")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18

    row = 1

    def title(text, r):
        c = ws.cell(row=r, column=1, value=text)
        c.font = styles["title_font"]
        return r + 1

    def section(text, r):
        c = ws.cell(row=r, column=1, value=text)
        c.font = styles["subtitle_font"]
        c.fill = styles["accent_fill"]
        ws.cell(row=r, column=2).fill = styles["accent_fill"]
        return r + 1

    def line(label, amount, r, bold=False):
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=amount)
        c2.number_format = '#,##0.00'
        if bold:
            c1.font = styles["bold_font"]
            c2.font = styles["bold_font"]
        else:
            c1.font = styles["data_font"]
            c2.font = styles["data_font"]
        return r + 1

    row = title(f"INCOME STATEMENT — {period}", row)
    row = title("(1 Jul 2024 – 30 Jun 2025)", row)
    row += 1

    row = section("Revenue", row)
    row = line("  GST-Free (Exports)", calc.get("revenue_gst_free", 0), row)
    row = line("  GST-Inclusive (Australian)", calc.get("revenue_gst_incl", 0), row)
    row = line("  GST Collected", -calc.get("revenue_gst_collected", 0), row)
    row = line("Total Revenue", calc["revenue_total"], row, bold=True)
    row += 1

    row = section("Cost of Services", row)
    row = line("  Contractors", -calc["cogs"], row)
    row = line("Gross Profit", calc["gross_profit"], row, bold=True)
    row += 1

    row = section("Operating Expenses", row)
    cat_labels = {
        "software": "Software & Subscriptions", "hosting": "Hosting & Infrastructure",
        "travel": "Travel & Transport", "meals": "Meals & Entertainment",
        "office": "Office & Coworking", "marketing": "Marketing & Advertising",
        "professional": "Professional Services", "other": "Other",
        "bank_fees": "Bank & FX Fees", "accounting": "Accounting",
        "computer": "Computer & Equipment", "utilities": "Utilities & Home Office",
        "phone": "Phone", "insurance": "Insurance", "cleaning": "Cleaning",
        "filing_fee": "ASIC Filing Fee", "director_fee": "Director Fee",
        "tax_payment": "Tax Payments",
    }
    # Show all opex categories sorted by amount
    for cat, v in sorted(calc["opex"].items(), key=lambda x: -x[1]):
        if v > 0:
            row = line(f"  {cat_labels.get(cat, cat)}", -v, row)
    if calc["uncategorized"] > 0:
        row = line("  Uncategorized", -calc["uncategorized"], row)
    row = line("Total Expenses", -(calc["total_opex"] + calc["uncategorized"]), row, bold=True)
    row += 1

    if calc.get("refund_total", 0) > 0:
        row = section("Refunds & Adjustments", row)
        row = line("  Refunds received", calc["refund_total"], row)
        row += 1

    ws.cell(row=row, column=1, value="").border = Border(bottom=Side(style="double"))
    ws.cell(row=row, column=2, value="").border = Border(bottom=Side(style="double"))
    row += 1
    row = line("Net Profit Before Tax", calc["net_profit"], row, bold=True)
    ws.cell(row=row - 1, column=2).fill = styles["lime_fill"]


def write_balance_sheet(wb, calc, period, styles):
    """Sheet 3: Balance Sheet."""
    ws = wb.create_sheet("Balance Sheet")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    row = 1
    def t(text, r):
        ws.cell(row=r, column=1, value=text).font = styles["title_font"]
        return r + 1
    def s(text, r):
        c = ws.cell(row=r, column=1, value=text)
        c.font = styles["subtitle_font"]
        c.fill = styles["accent_fill"]
        ws.cell(row=r, column=2).fill = styles["accent_fill"]
        return r + 1
    def l(label, amount, r, bold=False):
        ws.cell(row=r, column=1, value=label).font = styles["bold_font"] if bold else styles["data_font"]
        c = ws.cell(row=r, column=2, value=amount)
        c.number_format = '#,##0.00'
        c.font = styles["bold_font"] if bold else styles["data_font"]
        return r + 1

    row = t(f"BALANCE SHEET — As at 30 Jun 2025", row)
    row += 1
    row = s("ASSETS", row)
    row = s("  Current Assets", row)
    row = l("    Cash (Airwallex AUD)", calc["aud_balance"], row)
    row = l("    Total Current Assets", calc["aud_balance"], row, bold=True)
    row += 1
    row = l("  TOTAL ASSETS", calc["aud_balance"], row, bold=True)
    row += 1
    row = s("LIABILITIES", row)
    row = l("  TOTAL LIABILITIES", 0, row, bold=True)
    row += 1
    row = l("NET ASSETS", calc["aud_balance"], row, bold=True)
    row += 1
    row = s("EQUITY", row)
    retained = calc["net_profit"]  # Simplified — single year
    row = l("  Retained Earnings", retained, row)
    row = l("  TOTAL EQUITY", retained, row, bold=True)


def write_cashflow_sheet(wb, calc, period, styles):
    """Sheet 4: Cash Flow Statement (Direct Method)."""
    ws = wb.create_sheet("Cash Flow")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18

    row = 1
    def t(text, r):
        ws.cell(row=r, column=1, value=text).font = styles["title_font"]
        return r + 1
    def s(text, r):
        c = ws.cell(row=r, column=1, value=text)
        c.font = styles["subtitle_font"]
        c.fill = styles["accent_fill"]
        ws.cell(row=r, column=2).fill = styles["accent_fill"]
        return r + 1
    def l(label, amount, r, bold=False):
        ws.cell(row=r, column=1, value=label).font = styles["bold_font"] if bold else styles["data_font"]
        c = ws.cell(row=r, column=2, value=amount)
        c.number_format = '#,##0.00'
        c.font = styles["bold_font"] if bold else styles["data_font"]
        return r + 1

    row = t(f"CASH FLOW STATEMENT — {period} (Direct Method)", row)
    row += 1
    row = s("OPERATING ACTIVITIES", row)
    row = l("  Cash received from customers", calc["cash_received"], row)
    row = l("  Cash paid to suppliers", -calc["operating_paid"], row)
    row = l("  Net cash from operating", calc["operating_net"], row, bold=True)
    row += 1
    row = s("INVESTING ACTIVITIES", row)
    row = l("  Equipment purchases", -calc["investing"], row)
    row = l("  Net cash from investing", -calc["investing"], row, bold=True)
    row += 1
    row = s("FINANCING ACTIVITIES", row)
    row = l("  Net cash from financing", 0, row, bold=True)
    row += 1
    row = l("NET CHANGE IN CASH", calc["net_change"], row, bold=True)
    ws.cell(row=row - 1, column=2).fill = styles["lime_fill"]
    row += 1
    row = l("Current AUD Balance", calc["aud_balance"], row)


def write_category_summary(wb, calc, styles):
    """Sheet 5: Category Summary."""
    ws = wb.create_sheet("Category Summary")
    headers = ["Category", "Transactions", "AUD Total", "% of Total"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = styles["header_font"]
        c.fill = styles["header_fill"]
        c.alignment = styles["header_align"]

    grand_total = sum(v["total"] for v in calc["cat_summary"].values())
    sorted_cats = sorted(calc["cat_summary"].items(), key=lambda x: -x[1]["total"])

    for row_idx, (cat, v) in enumerate(sorted_cats, 2):
        ws.cell(row=row_idx, column=1, value=cat).font = styles["data_font"]
        ws.cell(row=row_idx, column=2, value=v["count"]).font = styles["data_font"]
        c = ws.cell(row=row_idx, column=3, value=v["total"])
        c.number_format = '#,##0.00'
        c.font = styles["data_font"]
        pct = v["total"] / grand_total if grand_total else 0
        c = ws.cell(row=row_idx, column=4, value=pct)
        c.number_format = '0.0%'
        c.font = styles["data_font"]

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12


def write_fx_summary(wb, calc, styles):
    """Sheet 6: FX Summary."""
    ws = wb.create_sheet("FX Summary")
    headers = ["Currency", "Transactions", "Original Total", "AUD Total", "Avg Rate", "Rate Source"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = styles["header_font"]
        c.fill = styles["header_fill"]
        c.alignment = styles["header_align"]

    sorted_fx = sorted(calc["fx_summary"].items(), key=lambda x: -x[1]["aud_total"])

    for row_idx, (cur, v) in enumerate(sorted_fx, 2):
        ws.cell(row=row_idx, column=1, value=cur).font = styles["data_font"]
        ws.cell(row=row_idx, column=2, value=v["count"]).font = styles["data_font"]
        ws.cell(row=row_idx, column=3, value=v["original_total"]).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=4, value=v["aud_total"]).number_format = '#,##0.00'
        avg_rate = sum(v["rates"]) / len(v["rates"]) if v["rates"] else None
        ws.cell(row=row_idx, column=5, value=avg_rate)
        if avg_rate:
            ws.cell(row=row_idx, column=5).number_format = '0.000000'
        sources = ", ".join(sorted(v["sources"])) if v["sources"] else "N/A"
        ws.cell(row=row_idx, column=6, value=sources).font = styles["data_font"]

    for i, w in enumerate([10, 14, 16, 16, 12, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_excel(txns, calc, period, output_dir):
    """Generate the multi-sheet Excel workbook."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    styles = xl_styles()

    write_ledger_sheet(wb, txns, styles)
    write_income_statement_sheet(wb, calc, period, styles)
    write_balance_sheet(wb, calc, period, styles)
    write_cashflow_sheet(wb, calc, period, styles)
    write_category_summary(wb, calc, styles)
    write_fx_summary(wb, calc, styles)

    path = output_dir / f"{period}-financial-pack.xlsx"
    wb.save(path)
    return path


# ─── HTML/PDF Generation ──────────────────────────────────────────────────────

def generate_html(calc, period, from_date, to_date, output_dir):
    """Generate print-optimised HTML for all three financial statements."""

    def statement_row(label, amount, bold=False, indent=0):
        cls = ' class="bold"' if bold else ""
        pad = "&nbsp;" * (indent * 4)
        sign = ""
        if amount is not None and amount < 0:
            sign = "negative"
        amt = fmt_aud(amount) if amount is not None else ""
        return f'<tr{cls}><td>{pad}{escape(label)}</td><td class="amount {sign}">{amt}</td></tr>'

    def section_header(text):
        return f'<tr class="section"><td colspan="2">{escape(text)}</td></tr>'

    def separator():
        return '<tr class="separator"><td colspan="2"></td></tr>'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Financial Statements — {escape(period)}</title>
<link rel="icon" type="image/png" href="{COMPANY_LOGO_URL}">
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: 'Segoe UI', system-ui, -apple-system, sans-serif; color: {BRAND_DARK}; background: #fff; padding: 40px; max-width: 800px; margin: 0 auto; }}
  .header {{ display: flex; align-items: center; gap: 16px; margin-bottom: 32px; padding-bottom: 16px; border-bottom: 3px solid {BRAND_DARK}; }}
  .header img {{ height: 48px; width: 48px; border-radius: 8px; }}
  .header h1 {{ font-size: 20px; font-weight: 700; }}
  .header .period {{ color: {BRAND_GREY}; font-size: 14px; }}
  .statement {{ margin-bottom: 48px; page-break-after: always; }}
  .statement:last-child {{ page-break-after: auto; }}
  .statement h2 {{ font-size: 18px; font-weight: 700; margin-bottom: 4px; color: {BRAND_DARK}; }}
  .statement .subtitle {{ font-size: 13px; color: {BRAND_GREY}; margin-bottom: 16px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  td {{ padding: 6px 8px; }}
  td.amount {{ text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }}
  td.amount.negative {{ color: #dc2626; }}
  tr.section td {{ font-weight: 700; background: {BRAND_LIGHT}; padding: 8px; border-top: 1px solid {BRAND_DARK}; }}
  tr.bold td {{ font-weight: 700; }}
  tr.separator td {{ border-bottom: 2px solid {BRAND_DARK}; padding: 0; height: 2px; }}
  tr.total td {{ font-weight: 700; background: {BRAND_PRIMARY}20; }}
  .footer {{ margin-top: 32px; padding-top: 16px; border-top: 1px solid #e2e8f0; font-size: 11px; color: {BRAND_GREY}; }}
  .disclaimer {{ background: #fffbeb; border: 1px solid #fcd34d; border-radius: 6px; padding: 12px; font-size: 12px; margin: 16px 0; }}
  @media print {{
    body {{ padding: 20px; }}
    .header {{ margin-bottom: 20px; }}
    .statement {{ margin-bottom: 0; }}
    .no-print {{ display: none; }}
  }}
</style>
</head>
<body>

<div class="header">
  <img src="{COMPANY_LOGO_URL}" alt="{YOUR_COMPANY}">
  <div>
    <h1>{YOUR_COMPANY} — Financial Statements</h1>
    <div class="period">{escape(period)} &mdash; {from_date} to {to_date}</div>
  </div>
</div>
"""

    cat_pct = round(100 * calc["categorized"] / calc["total_txns"], 0) if calc["total_txns"] else 0
    aud_pct = round(100 * calc["has_aud"] / calc["total_txns"], 0) if calc["total_txns"] else 0
    gst_pct = round(100 * calc["gst_classified"] / calc["total_txns"], 0) if calc["total_txns"] else 0

    html += f"""<div class="disclaimer">
  <strong>Transaction classification:</strong> {calc["total_txns"]} total transactions &mdash;
  {calc["revenue_txn_count"]} income, {calc["expense_txn_count"]} expense,
  {calc["internal_txn_count"]} internal (excluded), {calc["refund_txn_count"]} refunds.<br>
  Internal movements ({fmt_aud(calc["internal_total"])} — FX conversions, wallet transfers, Amex payments) are
  excluded from revenue and expenses.<br>
  <strong>Data quality:</strong> {cat_pct:.0f}% categorized, {aud_pct:.0f}% FX converted, {gst_pct:.0f}% GST classified.
</div>
"""

    # ── Income Statement ──
    html += '<div class="statement">\n'
    html += f'<h2>Income Statement (Profit & Loss)</h2>\n'
    html += f'<div class="subtitle">For the year ended 30 June 2025 — All amounts in AUD</div>\n'
    html += '<table>\n'

    html += section_header("Revenue")
    html += statement_row("GST-Free (Exports)", calc.get("revenue_gst_free", 0), indent=1)
    html += statement_row("GST-Inclusive (Australian)", calc.get("revenue_gst_incl", 0), indent=1)
    html += statement_row("GST Collected", -calc.get("revenue_gst_collected", 0), indent=1)
    html += statement_row("Total Revenue", calc["revenue_total"], bold=True)

    html += section_header("Cost of Services")
    html += statement_row("Contractors", -calc["cogs"], indent=1)
    html += statement_row("Gross Profit", calc["gross_profit"], bold=True)

    html += section_header("Operating Expenses")
    cat_labels = {
        "software": "Software & Subscriptions", "hosting": "Hosting & Infrastructure",
        "travel": "Travel & Transport", "meals": "Meals & Entertainment",
        "office": "Office & Coworking", "marketing": "Marketing & Advertising",
        "professional": "Professional Services", "other": "Other",
        "bank_fees": "Bank & FX Fees", "accounting": "Accounting",
        "computer": "Computer & Equipment", "utilities": "Utilities & Home Office",
        "phone": "Phone", "insurance": "Insurance", "cleaning": "Cleaning",
        "filing_fee": "ASIC Filing Fee", "director_fee": "Director Fee",
        "tax_payment": "Tax Payments",
    }
    for cat, v in sorted(calc["opex"].items(), key=lambda x: -x[1]):
        if v > 0:
            html += statement_row(cat_labels.get(cat, cat), -v, indent=1)
    if calc["uncategorized"] > 0:
        html += statement_row("Uncategorized", -calc["uncategorized"], indent=1)
    html += statement_row("Total Expenses", -(calc["total_opex"] + calc["uncategorized"]), bold=True)

    if calc["refund_total"] > 0:
        html += section_header("Refunds & Adjustments")
        html += statement_row("Refunds received", calc["refund_total"], indent=1)

    html += separator()
    html += f'<tr class="total"><td>Net Profit Before Tax</td><td class="amount">{fmt_aud(calc["net_profit"])}</td></tr>\n'
    html += '</table>\n</div>\n'

    # ── Balance Sheet ──
    html += '<div class="statement">\n'
    html += f'<h2>Balance Sheet (Statement of Financial Position)</h2>\n'
    html += f'<div class="subtitle">As at 30 June 2025 — All amounts in AUD</div>\n'
    html += '<table>\n'

    html += section_header("ASSETS")
    html += statement_row("Cash and cash equivalents (Airwallex)", calc["aud_balance"], indent=1)
    html += statement_row("TOTAL ASSETS", calc["aud_balance"], bold=True)

    html += section_header("LIABILITIES")
    html += statement_row("TOTAL LIABILITIES", 0, bold=True)

    html += separator()
    html += statement_row("NET ASSETS", calc["aud_balance"], bold=True)

    html += section_header("EQUITY")
    html += statement_row("Retained Earnings", calc["net_profit"], indent=1)
    html += statement_row("TOTAL EQUITY", calc["net_profit"], bold=True)

    html += '</table>\n</div>\n'

    # ── Cash Flow ──
    html += '<div class="statement">\n'
    html += f'<h2>Cash Flow Statement (Direct Method)</h2>\n'
    html += f'<div class="subtitle">For the year ended 30 June 2025 — All amounts in AUD</div>\n'
    html += '<table>\n'

    html += section_header("OPERATING ACTIVITIES")
    html += statement_row("Cash received from customers", calc["cash_received"], indent=1)
    html += statement_row("Cash paid to suppliers", -calc["operating_paid"], indent=1)
    html += statement_row("Net cash from operating", calc["operating_net"], bold=True)

    html += section_header("INVESTING ACTIVITIES")
    html += statement_row("Equipment purchases", -calc["investing"], indent=1)
    html += statement_row("Net cash from investing", -calc["investing"], bold=True)

    html += section_header("FINANCING ACTIVITIES")
    html += statement_row("Net cash from financing", 0, bold=True)

    html += separator()
    html += f'<tr class="total"><td>NET CHANGE IN CASH</td><td class="amount">{fmt_aud(calc["net_change"])}</td></tr>\n'
    html += statement_row("Current AUD Balance", calc["aud_balance"])

    html += '</table>\n</div>\n'

    # ── Footer ──
    now = datetime.utcnow().strftime("%d %b %Y %H:%M UTC")
    html += f"""<div class="footer">
  Generated by {YOUR_COMPANY} Finance Plugin on {now}.<br>
  Data sources: Airwallex API, Amex CSV import, RBA daily exchange rates.<br>
  {calc["total_txns"]} transactions processed. FX coverage: {aud_pct:.0f}%.
</div>

</body>
</html>"""

    path = output_dir / f"{period}-financial-statements.html"
    with open(path, "w") as f:
        f.write(html)
    return path


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Excel and HTML financial statements.")
    parser.add_argument("--period", required=True, help="Financial period (e.g., FY2025)")
    parser.add_argument("--output", default="all", choices=["excel", "html", "all"],
                        help="Output format (default: all)")
    args = parser.parse_args()

    from_date, to_date, period = parse_period(args.period)
    print(f"Period: {period} ({from_date} → {to_date})")

    # Load data
    data = load_data()
    balances = load_balances()
    txns = filter_period(data["transactions"], from_date, to_date)
    print(f"Transactions in period: {len(txns)}")

    if not txns:
        print("No transactions found for this period.", file=sys.stderr)
        sys.exit(1)

    # Load manual deductions if available
    manual = load_manual_deductions(period)
    if manual:
        print(f"Manual deductions: {len(manual.get('deductions', []))} items, ${manual.get('total', 0):,.2f}")

    # Calculate
    calc = calculate_statements(txns, balances, manual)

    # Output directory
    output_dir = Path(f"finance/tax/{period}")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Generate
    if args.output in ("excel", "all"):
        path = generate_excel(txns, calc, period, output_dir)
        print(f"Excel: {path} ({len(txns)} transactions, 6 sheets)")

    if args.output in ("html", "all"):
        path = generate_html(calc, period, from_date, to_date, output_dir)
        print(f"HTML:  {path} (open in browser, Ctrl+P for PDF)")

    # Always write JSON statements
    now = datetime.utcnow().isoformat() + "Z"

    income_stmt = {
        "period": period,
        "date_range": {"from": from_date, "to": to_date},
        "revenue": {
            "total": round(calc["revenue_total"], 2),
            "gst_free_exports": round(calc.get("revenue_gst_free", 0), 2),
            "gst_inclusive_au": round(calc.get("revenue_gst_incl", 0), 2),
            "gst_collected": round(calc.get("revenue_gst_collected", 0), 2),
        },
        "cost_of_services": {"contractors": round(calc["cogs"], 2), "total": round(calc["cogs"], 2)},
        "gross_profit": round(calc["gross_profit"], 2),
        "operating_expenses": {k: round(v, 2) for k, v in sorted(calc["opex"].items(), key=lambda x: -x[1]) if v > 0},
        "operating_expenses_total": round(calc["total_opex"], 2),
        "manual_deductions": {k: round(v, 2) for k, v in calc.get("manual_deductions", {}).items()},
        "manual_deductions_total": round(calc.get("manual_total", 0), 2),
        "uncategorized_expenses": round(calc["uncategorized"], 2),
        "refunds": round(calc["refund_total"], 2),
        "net_profit_before_tax": round(calc["net_profit"], 2),
        "completeness": {
            "categorized_pct": round(100 * calc["categorized"] / calc["total_txns"], 1) if calc["total_txns"] else 0,
            "fx_coverage_pct": round(100 * calc["has_aud"] / calc["total_txns"], 1) if calc["total_txns"] else 0,
            "gst_classified_pct": round(100 * calc["gst_classified"] / calc["total_txns"], 1) if calc["total_txns"] else 0,
        },
        "transaction_count": calc["total_txns"],
        "generated_at": now,
    }

    balance = {
        "period": period, "as_at": to_date,
        "assets": {"cash_aud": round(calc["aud_balance"], 2), "total": round(calc["aud_balance"], 2)},
        "liabilities": {"total": 0},
        "equity": {"retained_earnings": round(calc["net_profit"], 2), "total": round(calc["net_profit"], 2)},
        "generated_at": now,
    }

    cashflow = {
        "period": period, "date_range": {"from": from_date, "to": to_date},
        "operating": {"received": round(calc["cash_received"], 2), "paid": round(-calc["operating_paid"], 2), "net": round(calc["operating_net"], 2)},
        "investing": {"equipment": round(-calc["investing"], 2), "net": round(-calc["investing"], 2)},
        "financing": {"net": 0},
        "net_change": round(calc["net_change"], 2),
        "generated_at": now,
    }

    for name, obj in [("income-statement", income_stmt), ("balance-sheet", balance), ("cashflow-statement", cashflow)]:
        p = output_dir / f"{name}.json"
        with open(p, "w") as f:
            json.dump(obj, f, indent=2)
    print(f"JSON:  {output_dir}/{{income-statement,balance-sheet,cashflow-statement}}.json")

    # Print key figures
    print(f"\n{'='*50}")
    print(f"Revenue:     {fmt_aud(calc['revenue_total'])}")
    print(f"  GST-free:  {fmt_aud(calc.get('revenue_gst_free', 0))}")
    print(f"  GST-incl:  {fmt_aud(calc.get('revenue_gst_incl', 0))}")
    print(f"COGS:        {fmt_aud(-calc['cogs'])}")
    print(f"OpEx:        {fmt_aud(-calc['total_opex'])}")
    print(f"Refunds:     {fmt_aud(calc['refund_total'])}")
    print(f"Net Profit:  {fmt_aud(calc['net_profit'])}")
    print(f"{'='*50}")

    print("Done.")


if __name__ == "__main__":
    main()
